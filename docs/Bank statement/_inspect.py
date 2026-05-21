"""
Quick inspector for the bank-statement dump in this folder.

Reports per-file:
  - source format (xlsx / xls / pdf)
  - row + column count
  - the first non-empty row that looks like a header
  - the first 2 data rows for shape verification
  - inferred totals (row count, date range, sums) if columns can be detected

No values are written anywhere — read-only inspection. The transaction lines
themselves are NOT echoed in full (only header + 2 sample rows) so the output
stays readable.
"""
from __future__ import annotations
import os
import sys
from pathlib import Path

HERE = Path(__file__).parent

# Mask numbers/strings that look like account numbers or full long digit strings
import re
ACCT_RE = re.compile(r"\b\d{9,18}\b")
def _mask(s: str | None) -> str:
    if s is None:
        return ""
    return ACCT_RE.sub(lambda m: f"••{m.group(0)[-4:]}", str(s))


def inspect_xlsx(p: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        # First row that has at least 3 non-empty cells is treated as header candidate.
        header_idx = next(
            (i for i, r in enumerate(rows) if r and sum(1 for c in r if c not in (None, "")) >= 3),
            None,
        )
        header = rows[header_idx] if header_idx is not None else None
        data_rows = rows[header_idx + 1:] if header_idx is not None else rows
        # Drop trailing/leading blank rows
        data_rows = [r for r in data_rows if r and any(c not in (None, "") for c in r)]
        sheets.append({
            "name": ws.title,
            "total_rows": ws.max_row,
            "total_cols": ws.max_column,
            "header_idx": header_idx,
            "header": header,
            "data_row_count": len(data_rows),
            "first_two": data_rows[:2],
            "last": data_rows[-1] if data_rows else None,
        })
    return {"format": "xlsx", "sheets": sheets}


def inspect_xls(p: Path) -> dict:
    import xlrd
    book = xlrd.open_workbook(p)
    sheets = []
    for sh in book.sheets():
        rows = [sh.row_values(r) for r in range(sh.nrows)]
        header_idx = next(
            (i for i, r in enumerate(rows) if sum(1 for c in r if c not in (None, "")) >= 3),
            None,
        )
        header = rows[header_idx] if header_idx is not None else None
        data_rows = rows[header_idx + 1:] if header_idx is not None else rows
        data_rows = [r for r in data_rows if any(c not in (None, "") for c in r)]
        sheets.append({
            "name": sh.name,
            "total_rows": sh.nrows,
            "total_cols": sh.ncols,
            "header_idx": header_idx,
            "header": header,
            "data_row_count": len(data_rows),
            "first_two": data_rows[:2],
            "last": data_rows[-1] if data_rows else None,
        })
    return {"format": "xls", "sheets": sheets}


def inspect_pdf(p: Path) -> dict:
    try:
        import pdfplumber
    except ImportError:
        return {"format": "pdf", "error": "pdfplumber not installed"}
    out = {"format": "pdf", "pages": 0, "first_lines": [], "table_rows": 0}
    with pdfplumber.open(p) as pdf:
        out["pages"] = len(pdf.pages)
        first = pdf.pages[0]
        text = first.extract_text() or ""
        out["first_lines"] = text.splitlines()[:12]
        # Try table extraction across all pages
        total = 0
        for page in pdf.pages:
            for t in page.extract_tables() or []:
                total += max(0, len(t) - 1)  # subtract header per page
        out["table_rows"] = total
    return out


def main() -> None:
    files = sorted(p for p in HERE.iterdir() if p.is_file() and not p.name.startswith("_"))
    print(f"Inspecting {len(files)} files in {HERE}\n")
    for p in files:
        ext = p.suffix.lower()
        print("=" * 78)
        print(f"FILE: {p.name}")
        print(f"  size: {p.stat().st_size:,} bytes")
        try:
            if ext == ".xlsx":
                info = inspect_xlsx(p)
            elif ext == ".xls":
                info = inspect_xls(p)
            elif ext == ".pdf":
                info = inspect_pdf(p)
            else:
                print(f"  skipped (unknown extension {ext})")
                continue
        except Exception as e:
            print(f"  ERROR: {type(e).__name__}: {e}")
            continue
        print(f"  format: {info['format']}")
        if "sheets" in info:
            for s in info["sheets"]:
                print(f"  sheet '{s['name']}' — {s['total_rows']} rows x {s['total_cols']} cols")
                print(f"    header (row {s['header_idx']}): {_mask(' | '.join(str(c or '') for c in (s['header'] or [])))[:160]}")
                print(f"    data rows: {s['data_row_count']}")
                for i, r in enumerate(s['first_two']):
                    print(f"    sample {i+1}: {_mask(' | '.join(str(c) for c in r))[:160]}")
                if s['last'] is not None and s['data_row_count'] > 2:
                    print(f"    last:    {_mask(' | '.join(str(c) for c in s['last']))[:160]}")
        elif info['format'] == 'pdf':
            print(f"  pages: {info.get('pages')}")
            print(f"  detected table rows (sum): {info.get('table_rows')}")
            print("  first text lines:")
            for line in info.get("first_lines", []):
                print(f"    {_mask(line)[:160]}")
            if "error" in info:
                print(f"  error: {info['error']}")
        print()


if __name__ == "__main__":
    main()
